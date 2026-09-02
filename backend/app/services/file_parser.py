"""文件解析：把上传的非标准文件解析成『文本 + 表格』，供 LLM 结构化。
- Excel/CSV：pandas 读出所有 sheet，转 markdown 表格文本
- PDF：pdfplumber 提取文本和表格
- 图片：返回 bytes，走视觉模型
"""
import re
import uuid
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import pdfplumber

from ..config import DATA_DIR

UPLOAD_DIR = DATA_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".webp", ".gif"}
EXCEL_EXTS = {".xlsx", ".xls", ".xlsm", ".csv", ".tsv"}
SPREADSHEET_EXTS = {".xlsx", ".xls", ".xlsm"}
_OLE_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


class FileParseError(ValueError):
    """内容不是声称的格式，给用户看短中文原因。"""


def spreadsheet_reject_reason(filename: str, data: bytes) -> str | None:
    """扩展名是 Excel 但魔数对不上时返回短中文原因，否则 None。"""
    ext = Path(filename or "").suffix.lower()
    if ext not in SPREADSHEET_EXTS:
        return None
    if not data:
        return "文件为空"
    head = data[:80]
    if head.startswith(b"PK") or head.startswith(_OLE_MAGIC):
        return None
    if head.startswith(b"%TSD-Header") or b"%TSD-Header" in data[:256]:
        return (
            "文件被加密或网盘封装（TSD），不是真正的 Excel。"
            "请用 Excel 打开后「另存为 .xlsx」再上传。"
        )
    sample = data[:1024].lstrip().lower()
    if sample.startswith(b"<") or b"<html" in sample:
        return "文件实际是 HTML 网页，不是 Excel。请另存为 xlsx 后再传。"
    return (
        "扩展名是 Excel，但内容不是有效的 xlsx/xls。"
        "常见原因：加密、网盘封装、下载损坏。请另存为 xlsx 后再传。"
    )


def save_upload(filename: str, content: bytes) -> dict:
    """保存上传文件，返回 {file_id, filename, ext, path}"""
    ext = Path(filename).suffix.lower()
    file_id = uuid.uuid4().hex[:12]
    safe_name = f"{file_id}{ext}"
    path = UPLOAD_DIR / safe_name
    path.write_bytes(content)
    (UPLOAD_DIR / f".name-{file_id}").write_text(filename, encoding="utf-8")
    return {"file_id": file_id, "filename": filename, "ext": ext, "path": str(path)}


def original_filename(file_id: str) -> str:
    meta = UPLOAD_DIR / f".name-{file_id}"
    if meta.exists():
        name = meta.read_text(encoding="utf-8").strip()
        if name:
            return name
    path = _get_path(file_id)
    return path.name


def _get_path(file_id: str) -> Path:
    matches = list(UPLOAD_DIR.glob(f"{file_id}.*"))
    if not matches:
        raise FileNotFoundError(f"file_id {file_id} 不存在")
    return matches[0]


def is_image(file_id: str) -> bool:
    return _get_path(file_id).suffix.lower() in IMAGE_EXTS


def get_image_bytes(file_id: str) -> tuple[bytes, str]:
    path = _get_path(file_id)
    return path.read_bytes(), path.suffix.lower().lstrip(".")


def parse_to_text(file_id: str, max_chars: int = 0) -> str:
    """把文件内容解析成文本（markdown 表格优先）。
    max_chars <= 0 表示不截断；仅当 max_chars > 0 且超长时才截断。
    """
    path = _get_path(file_id)
    ext = path.suffix.lower()

    if ext in EXCEL_EXTS:
        text = _parse_excel(path)
    elif ext == ".pdf":
        text = _parse_pdf(path)
    elif ext in {".txt", ".md", ".json"}:
        text = path.read_text(encoding="utf-8", errors="replace")
    else:
        raise ValueError(f"不支持的文件类型: {ext}（扫描件/图片请走视觉模型）")

    if max_chars and max_chars > 0 and len(text) > max_chars:
        orig = len(text)
        text = text[:max_chars] + f"\n...(已截断，共 {orig} 字符)"
    return text


_FMT_DATE = re.compile(r"[yY]{2,}|[mM]{2,}|[dD]{2,}")
_FMT_SCI = re.compile(r"[eE][+\-0]")


def excel_decimal_places(number_format: str) -> int | None:
    """从 Excel 数字格式读出小数位。General / 无法识别则返回 None。"""
    fmt = (number_format or "General").split(";")[0]
    fmt = fmt.replace("\\", "").replace("_", "").replace("*", "").replace("?", "")
    fmt = re.sub(r'"[^"]*"', "", fmt)
    if "General" in fmt or fmt.strip() in {"@", ""}:
        return None
    if _FMT_SCI.search(fmt):
        head = fmt.split("e")[0].split("E")[0]
        if "." in head:
            return sum(1 for ch in head.split(".", 1)[1] if ch in "0#")
        return 0
    if "." in fmt:
        dec = fmt.split(".", 1)[1]
        dec = dec.split("%")[0]
        n = sum(1 for ch in dec if ch in "0#")
        return n
    if "0" in fmt or "#" in fmt:
        return 0
    return None


def format_excel_number(value: float, number_format: str = "General") -> str:
    """按单元格显示格式输出数字，而不是 IEEE 原始精度。"""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    fmt = number_format or "General"
    if _FMT_DATE.search(fmt) and "E+" not in fmt.upper() and "%" not in fmt:
        try:
            from openpyxl.utils.datetime import from_excel
            dt = from_excel(value)
            if isinstance(dt, datetime):
                return dt.strftime("%Y-%m-%d")
        except Exception:
            pass
    places = excel_decimal_places(fmt)
    num = float(value)
    if _FMT_SCI.search(fmt) and "%" not in fmt:
        prec = 2 if places is None else places
        return f"{num:.{prec}E}"
    if "%" in fmt:
        num *= 100
        if places is None:
            places = 0
        return f"{num:.{places}f}%"
    if places is None:
        if isinstance(value, int) or (isinstance(value, float) and num.is_integer() and abs(num) < 1e15):
            return str(int(num))
        text = f"{num:.11g}"
        if "e" in text.lower():
            return text
        if "." in text:
            text = text.rstrip("0").rstrip(".")
        return text or "0"
    rounded = f"{num:.{places}f}"
    return rounded


def excel_display_value(value, number_format: str = "General") -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S").replace(" 00:00:00", "")
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return format_excel_number(value, number_format)
    return str(value).strip()


def _df_to_markdown(df: pd.DataFrame) -> str:
    """DataFrame 转 markdown 表格，列名不清时保留原始内容"""
    df = df.replace("", pd.NA)
    df = df.dropna(how="all").dropna(axis=1, how="all")
    if df.empty:
        return ""
    df = df.fillna("")
    df.columns = [str(c) if str(c) and not str(c).startswith("Unnamed") else f"col{i}" for i, c in enumerate(df.columns)]
    return df.to_markdown(index=False)


def _read_xlsx_displayed(path: Path) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), data_only=True)
    try:
        out = {}
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows():
                rows.append([excel_display_value(c.value, c.number_format) for c in row])
            if rows and any(any(x for x in r) for r in rows):
                out[ws.title] = pd.DataFrame(rows)
        if not out:
            raise ValueError("openpyxl displayed: empty")
        return out
    finally:
        wb.close()


def _read_xls_displayed(path: Path) -> dict:
    import xlrd

    try:
        book = xlrd.open_workbook(str(path), formatting_info=True)
    except Exception:
        book = xlrd.open_workbook(str(path), formatting_info=False)
    fmt_map = getattr(book, "format_map", {})
    xf_list = getattr(book, "xf_list", [])

    def cell_fmt(cell) -> str:
        try:
            xf = xf_list[cell.xf_index]
            rec = fmt_map.get(xf.format_key)
            return getattr(rec, "format_str", None) or "General"
        except Exception:
            return "General"

    out = {}
    for sheet in book.sheets():
        rows = []
        for r in range(sheet.nrows):
            line = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                val = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        tup = xlrd.xldate_as_tuple(val, book.datemode)
                        val = datetime(*tup[:6]) if any(tup[3:]) else date(*tup[:3])
                    except Exception:
                        pass
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    val = bool(val)
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    val = None
                line.append(excel_display_value(val, cell_fmt(cell)))
            rows.append(line)
        if rows and any(any(x for x in r) for r in rows):
            out[sheet.name] = pd.DataFrame(rows)
    if not out:
        raise ValueError("xlrd displayed: empty")
    return out


def _read_excel_displayed(path: Path) -> dict:
    ext = path.suffix.lower()
    if ext == ".xls":
        return _read_xls_displayed(path)
    return _read_xlsx_displayed(path)


def _read_excel_sheets(path: Path) -> dict:
    """CRO 报告里常见：xls、伪 xlsx、openpyxl 自定义属性 name=None。逐个引擎兜底。"""
    errors: list[str] = []
    for engine in (None, "calamine", "openpyxl", "xlrd"):
        try:
            kw = {"sheet_name": None, "dtype": str, "header": None}
            if engine:
                kw["engine"] = engine
            return pd.read_excel(path, **kw)
        except Exception as e:
            errors.append(f"{engine or 'auto'}: {e}")
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        out = {}
        for ws in wb.worksheets:
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            if not rows:
                continue
            out[ws.title] = pd.DataFrame(rows)
        wb.close()
        if out:
            return out
        errors.append("openpyxl-readonly: empty")
    except Exception as e:
        errors.append(f"openpyxl-readonly: {e}")
    raise FileParseError(
        "不是有效的 Excel 文件（可能加密、网盘封装或已损坏）。请另存为 xlsx 后再传。"
    )


def _parse_excel(path: Path) -> str:
    ext = path.suffix.lower()
    parts = []
    if ext in {".csv", ".tsv"}:
        sep = "\t" if ext == ".tsv" else None
        df = pd.read_csv(path, sep=sep, engine="python", dtype=str, header=None)
        parts.append(_df_to_markdown(df))
    else:
        reason = spreadsheet_reject_reason(path.name, path.read_bytes())
        if reason:
            raise FileParseError(reason)
        try:
            sheets = _read_excel_displayed(path)
        except Exception:
            sheets = _read_excel_sheets(path)
        for name, df in sheets.items():
            md = _df_to_markdown(df)
            if md:
                parts.append(f"### Sheet: {name}\n{md}")
    return "\n\n".join(p for p in parts if p)


def _parse_pdf(path: Path) -> str:
    parts = []
    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            if text.strip():
                parts.append(f"### 第 {i+1} 页\n{text}")
            for t_idx, table in enumerate(page.extract_tables()):
                df = pd.DataFrame(table)
                md = _df_to_markdown(df)
                if md:
                    parts.append(f"### 第 {i+1} 页 表格 {t_idx+1}\n{md}")
    return "\n\n".join(parts)
